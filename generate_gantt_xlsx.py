from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# ── Data ──
tasks = [
    ("Idea Finalization",              "", 100, "Completed",   1, 1),
    ("Problem Research & SDG Mapping", "", 30,  "Completed",   1, 1.5),
    ("Market & Competitor Analysis",   "", 40,  "Not Started", 1.5, 2.5),
    ("Requirement Gathering",          "", 60,  "Started",     1, 2),
    ("System Design (UI/UX)",          "", 100, "Started",     2, 4),
    ("Database Design",                "", 10,  "Not Started", 2, 3),
    ("Video Presentation",             "", 0,   "Not Started", 7, 8),
    ("Frontend Development",           "", 100, "Started",     2, 5),
    ("Backend Development",            "Prabin Sharma", 90, "Not Started", 3, 6),
    ("AI Integration",                 "", 0,   "Not Started", 3, 5),
    ("Testing & Debugging",            "", 70,  "Not Started", 5, 7),
    ("Deployment",                     "", 0,   "Not Started", 6, 7),
    ("User Feedback Collection",       "", 5,   "Not Started", 6, 8),
    ("Improvement & Optimization",     "", 0,   "Not Started", 7, 8),
    ("Documentation",                  "", 0,   "Not Started", 5, 7),
    ("Customer Support / Maintenance", "", 0,   "Not Started", 7, 8),
]

weeks = 8

# ── Colors ──
DARK_GREEN = "1B5E20"
WHITE = "FFFFFF"
HEADER_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)

COMPLETED_FILL = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
STARTED_FILL = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
NOT_STARTED_FILL = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

STATUS_FILLS = {
    "Completed":   COMPLETED_FILL,
    "Started":     STARTED_FILL,
    "Not Started": NOT_STARTED_FILL,
}
STATUS_FONTS = {
    "Completed":   Font(name="Calibri", size=9, bold=True, color="FFFFFF"),
    "Started":     Font(name="Calibri", size=9, bold=True, color="000000"),
    "Not Started": Font(name="Calibri", size=9, bold=True, color="FFFFFF"),
}

# Gantt bar fills — completed portion vs remaining
GANTT_DONE = {
    "Completed":   PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "Started":     PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Not Started": PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),
}
GANTT_REMAINING = {
    "Completed":   PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid"),
    "Started":     PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid"),
    "Not Started": PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid"),
}
GANTT_DONE_FONT = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
GANTT_REM_FONT = Font(name="Calibri", size=8, bold=False, color="333333")

def progress_fill(pct):
    if pct >= 100:
        return PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    elif pct >= 70:
        return PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
    elif pct >= 40:
        return PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")
    elif pct > 0:
        return PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
    else:
        return PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# ── Headers ──
headers = ["#", "Task", "Owner", "Progress (%)", "Status"]
week_headers = [f"Week {i}" for i in range(1, weeks + 1)]
all_headers = headers + week_headers

for col_idx, header in enumerate(all_headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# ── Column widths ──
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
for i in range(weeks):
    ws.column_dimensions[get_column_letter(6 + i)].width = 12

ws.row_dimensions[1].height = 25

# ── Data rows ──
for row_idx, (task, owner, progress, status, start_w, end_w) in enumerate(tasks, 2):
    ws.row_dimensions[row_idx].height = 24

    # #
    cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    cell.font = Font(name="Calibri", size=9, color="555555")

    # Task
    cell = ws.cell(row=row_idx, column=2, value=task)
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    cell.font = Font(name="Calibri", size=10, bold=True)

    # Owner
    cell = ws.cell(row=row_idx, column=3, value=owner)
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    cell.font = Font(name="Calibri", size=9)

    # Progress
    cell = ws.cell(row=row_idx, column=4, value=progress)
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF" if progress >= 40 else "333333")
    cell.fill = progress_fill(progress)

    # Status
    cell = ws.cell(row=row_idx, column=5, value=status)
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    cell.fill = STATUS_FILLS.get(status, PatternFill())
    cell.font = STATUS_FONTS.get(status, Font(name="Calibri", size=9))

    # ── Gantt bars with percentage split ──
    # Build list of week cells that belong to this task
    task_weeks = []
    for w in range(1, weeks + 1):
        if w >= int(start_w) and w <= int(end_w - 0.01) + 1:
            if start_w % 1 != 0 and w == int(start_w):
                continue  # skip partial start if half-week offset
            task_weeks.append(w)
        # Handle fractional starts
        if start_w % 1 != 0 and w == int(start_w) + 1:
            if w not in task_weeks:
                task_weeks.append(w)

    # Recalculate cleanly
    task_weeks = []
    for w in range(1, weeks + 1):
        # A week is part of the task if any portion overlaps
        week_start = w
        week_end = w + 1
        if week_start < end_w and week_end > start_w:
            task_weeks.append(w)

    total_cells = len(task_weeks)
    if total_cells > 0:
        done_cells = max(0, round(total_cells * progress / 100))
    else:
        done_cells = 0

    done_fill = GANTT_DONE.get(status)
    rem_fill = GANTT_REMAINING.get(status)

    for idx, w in enumerate(task_weeks):
        col = 5 + w
        cell = ws.cell(row=row_idx, column=col)
        cell.border = THIN_BORDER
        cell.alignment = CENTER

        if idx < done_cells:
            # Completed portion — dark color with percentage
            cell.fill = done_fill
            cell.font = GANTT_DONE_FONT
            # Show percentage only on the middle cell of the done section
            if idx == done_cells // 2:
                cell.value = f"{progress}%"
        else:
            # Remaining portion — light color
            cell.fill = rem_fill
            cell.font = GANTT_REM_FONT

    # For empty week cells (not part of task), just border
    for w in range(1, weeks + 1):
        if w not in task_weeks:
            col = 5 + w
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER

# ── Color Legend (below the data) ──
legend_row = len(tasks) + 3
ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Calibri", size=10, bold=True)

legends = [
    ("Completed (done)", "2E7D32", "FFFFFF"),
    ("Completed (remaining)", "A5D6A7", "333333"),
    ("Started (done)", "1565C0", "FFFFFF"),
    ("Started (remaining)", "90CAF9", "333333"),
    ("Not Started (done)", "C62828", "FFFFFF"),
    ("Not Started (remaining)", "EF9A9A", "333333"),
]

for i, (label, bg, fg) in enumerate(legends):
    col = 2 + i
    cell = ws.cell(row=legend_row, column=col, value=label)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(name="Calibri", size=8, bold=True, color=fg)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# ── Freeze panes ──
ws.freeze_panes = "F2"

# ── Save ──
output = "/Users/prabinsharma/Desktop/StayNep_Gantt_Chart.xlsx"
wb.save(output)
print(f"Saved to: {output}")
